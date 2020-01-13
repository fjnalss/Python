from django.shortcuts import render, redirect
from django.http import HttpResponseRedirect, HttpResponse
from django.urls import reverse
from django.contrib.auth.decorators import login_required
from .models import User, Oui, Url, Result
import requests
from openpyxl import  Workbook
from io import BytesIO
from django.utils.http import urlquote
import datetime
from .sam import Oho, Url


@login_required(login_url='/autovalidation/login')
def index(request):
    return render(request, 'verify/index.html')


def login(request):
    if request.session.get('is_login',None):
        return render(request, 'verify/index.html')
    if request.method == 'GET':
        return render(request, 'verify/login.html')
    else:
        name = request.POST['u_name']
        pwd = request.POST['pwd']
        passwd = User.objects.filter(username=name).values_list('passwd', flat=True)
        user_id = User.objects.filter(username=name).values_list('id', flat=True)
        if passwd[0] == pwd:
            request.session['is_login'] = True
            request.session['user_id'] = user_id[0]
            request.session['user_name'] = name
            return render(request,'verify/index.html')
        else:
            return render(request, 'verify/login.html',{'error_msg': "账号或密码错误", 'name':name, 'passwd':passwd})


def logout(request):
    request.session.flush()
    return HttpResponseRedirect(reverse('verify:index'))


def verify(request):
    if request.method == 'POST':
        domains = request.POST['domains']
        if len(domains) < 1:
            return render(request, 'verify/verify.html', {'error_msg': "请输入域名再验证"})
        domains_list = domains.split(' ')
        for domain in domains_list:
            if domain:
                verify_function(request, domain)
        result = Result.objects.filter(u_id=request.session.get('user_id'), domain__in=domains_list)
        return render(request, 'verify/result.html', {'domain_list': domains_list, 'result': result})
    else:
        return render(request, 'verify/verify.html')

#验证域名包含的功能，将有问题的功能放在error字段，只要error字段有值，则验证不ok
def verify_function(request, domain):
    oui_cache = Oui.objects.filter(domain=domain).values_list('force_max_age', flat=True)
    error_list = []
    if oui_cache:
        ok = verify_cache(domain)
        if ok == 0:
            error_list.append('cache')
    #其他的验证功能
    error = ','.join(error_list)
    ok = 1
    if len(error) > 0:
        ok = 0
    obj, created = Result.objects.update_or_create(u_id=request.session.get('user_id'), domain=domain, defaults={'ok':ok, 'error':error})


# 验证域名的缓存功能
def verify_cache(domain):
    cname_cnc = domain+".wtxcdn.com:80"
    cname_cdnw = domain + ".cdnga.net:80"
    proxies_cnc = {'http': cname_cnc}
    proxies_cdnw = {'http': cname_cdnw}
    uris = Url.objects.filter(domain=domain).values_list('uri', flat=True)
    params = Url.objects.filter(domain=domain).values_list('param', flat=True)
    res_cnc = []
    res_cdnw = []
    for uri, param in zip(uris, params):
        res_cnc.append(requests.head("http://"+domain+uri, proxies=proxies_cnc, params=param))
        res_cdnw.append(requests.head("http://"+domain+uri, proxies=proxies_cdnw, params=param))
    for cnc, cdnw in zip(res_cnc, res_cdnw):
        if cnc.status_code != cdnw.status_code:
            return 0
        if "Content-Length" in cnc.headers:
            if cnc.headers['Content-Length'] != cdnw.headers['Content-Length']:
                return 0
        # if cnc.headers['Cache-Control'] != cdnw.headers['Cache-Control']:
        #     return 0
    return 1


def export(request):
    wb = Workbook()
    wb.encoding = 'utf-8'
    sheet1 = wb.active
    sheet1.title = '测试报告'
    table_title = ['域名', '验证结果', '错误的功能']
    for i in range(1, len(table_title)+1):
        sheet1.cell(row=1, column=i).value=table_title[i-1]
    all_obj = Result.objects.filter(u_id=request.session.get('user_id')) # 取得该用户的所有域名，不是该次测试的域名
    for obj in all_obj:
        max_row = sheet1.max_row+1
        obj_info = [obj.domain, obj.ok, obj.error]
        for x in range(1, len(obj_info)+1):
            sheet1.cell(row=max_row, column=x).value = obj_info[x-1]
    output = BytesIO()
    wb.save(output) # 将excel保存到IO
    output.seek(0)
    response = HttpResponse(output.getvalue(), content_type='application/vnd.ms-excel')
    ctime = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    file_name = '测试报告%s.xlsx' % ctime
    file_name = urlquote(file_name) # 使用urlquote()方法解决中文无法使用的问题
    response['Content-Disposition'] = 'attachment; filename=%s' % file_name
    return response


def migrate(request):
    if request.method == 'GET':
        return render(request, 'migrate/index.html')
    elif request.method == 'POST':
        domain = request.POST.get('domain')
        return render(request, 'migrate/migrate.html', {'domain': domain})
    else:
        return render(request, 'migrate/index.html', {'error_msg': "请求方法有问题"})
